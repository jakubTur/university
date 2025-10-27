import sys
import os
import logging
import csv
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

load_dotenv(dotenv_path="parameters.env")

SUBJECTS_TAKEN=int(os.getenv('SUBJECTS_TAKEN', 1))
REQUIRED_SUBJECT=os.getenv('REQUIRED_SUBJECT', "")


def dataList(filename):
    dataFile = open(filename)
    dataReader = csv.reader(dataFile)
    dataList = list(dataReader)
    dataFile.close()
    dataList[0].pop(4)
    for line in dataList[1:]:
        line.pop(0)
        line.pop(4)
    return dataList

def average(data):
    total=0
    notCounted=1
    if(REQUIRED_SUBJECT!=""):
        index=data[0].index(REQUIRED_SUBJECT)
    else:
        index=len(data[0])-1
    for line in data[1:]:
        avg=0
        count=0
        for parameter in line[:-1]:
            if(parameter!=""):
                avg+=float(parameter)
                count+=1
        if(avg!=0 and count>=SUBJECTS_TAKEN and line[index]!=''):
            total+=(avg/count)
        else:
            notCounted+=1
    total/=(len(data)-notCounted)
    return total

def averageByStudent(data):
    averages=[]
    for line in data[1:]:
        avg=0
        count=0
        for parameter in line[:-1]:
            if(parameter!=""):
                avg+=float(parameter)
                count+=1
        if(avg!=0):
            averages.append((line[len(line)-1], avg/count))
        else:
            averages.append((line[len(line)-1], 0))
    return averages

def studentsTotal(data, subject):
    if(subject not in data[0]):
        if(subject=="foreign language" or subject=="foreign_language"):
            index=4
        elif(subject=="civic education"):
            index=1
        else:
            logging.error("this is not a valid subject")
            return
    else:
        index=data[0].index(subject)
        totalCount=0
        for line in data[1:]:
            if(line[index]!=""):
                totalCount+=1
        return totalCount

def saveAsExcel(filename, avg, avgByStudent, totalNumber):
    wb=Workbook()
    ws=wb.active
    ws.append(["student id", "average score"])
    font=Font(bold=True, color="FFFFFF")
    fill=PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    for cell in ws[1]:
        cell.font = font
        cell.fill = fill
    for id, avg in avgByStudent:
        ws.append([id, avg])
    ws.append([])
    ws.append(["average of all students:",avg])
    ws.append(["number of students who took physics:",totalNumber])
    wb.save(filename)

def run():
    if(len(sys.argv)<2):
        logging.error("wrong amount of arguments")
        sys.exit(1)
    filename=sys.argv[1]
    if(not filename.endswith('.csv')):
        if(filename=="-h"):
            print("this application is used to browse database of scores in final exams of Vietnamese children" \
            "\nyou can calculate the average score of all students, specific students, or calculate how many people took a specific course (physics by default)" \
            "\nyou can edit parameters.env to provide default values for how many subjects a student needs to take to ba taken into account"
            "\nor what subject they need to take to calculate the total average"
            "\nadd -o filename to save as excel document under filename name")
            sys.exit(1)
        else:
            logging.error("error: The file must have a '.csv' extension.")
            sys.exit(1)
    if not os.path.isfile(filename):
        logging.error(f"error: the file '{filename}' does not exist.")
        sys.exit(1)
    
    data=dataList(filename)

    tuples=averageByStudent(data)
    avg=(average(data))
    phys=studentsTotal(data, "physics")

    if(len(sys.argv)>2 and sys.argv[2]=="-o"):
        if(len(sys.argv)>3 and sys.argv[3].endswith(".xlsx")):
            destination=sys.argv[3]
            saveAsExcel(destination, avg, tuples, phys)
        else:
            logging.error(f"error: destination file has to be an excel file.")
            sys.exit(1)
    else:
        print(f"taking into account students who took {REQUIRED_SUBJECT} and at least {SUBJECTS_TAKEN} subjects")
        print(f"average of all students: {avg}")
        print(f'number of students who took physics: {phys}')
            
 


if __name__ == "__main__":
    run()
else:
    logging.error("This file must be run directly, not imported.")
    sys.exit(1)